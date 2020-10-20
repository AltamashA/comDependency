from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
import time
import os

class excel:
    """
        helper function for excel saving data
    """
    def __init__(self,start=None,end=None):
        if start==None :
            end = datetime.now().minute
            start = datetime.now().hour
        self.wb = Workbook()
        self.ws = self.wb.active
        self.dir = datetime.now().strftime("%m-%d-%Y")
        self.file = datetime.now().strftime("%m-%d-%Y")+'_'+str(start)+"-"+str(end)
        self.headerFont = Font(color='808080' ,bold=True)

    def __makeDir(self):
        if not os.path.exists(self.dir):
            os.makedirs(self.dir)

    def save(self,header,rows):
        self.__makeDir()
        self.ws.insert_rows(1,header)
        for cell in self.ws["1:1"]:
            cell.font=self.headerFont
        for row in rows:
            self.ws.append(row)
    def __saveHeader(self,level):
        source = ["SERVICE","METHOD","PATH"]
        for i in range(0,len(source)):
            self.ws.cell(row=1,column=i+1).value = source[i]
        self.ws.merge_cells(start_row=1,end_row=2,start_column=1,end_column=1)
        self.ws.merge_cells(start_row=1,end_row=2,start_column=2,end_column=2)
        self.ws.merge_cells(start_row=1,end_row=2,start_column=3,end_column=3)
        for x in range(0,level-1):
            self.ws.cell(row=1,column=3*(x+1)+1).value = "DESTINATION"
            self.ws.merge_cells(start_row=1,end_row=1,start_column=3*(x+1)+1,end_column=3*(x+2))
            self.ws.cell(row=2,column=3*(x+1)+1).value = "SERVICE"
            self.ws.cell(row=2,column=3*(x+1)+2).value = "METHOD"
            self.ws.cell(row=2,column=3*(x+1)+3).value = "PATH"
        self.wb.save(self.dir+'/'+self.file+'.xlsx')
    def __insert(self,rows,col_num):
        for row in rows:
            row_num = self.ws.max_row
            self.ws.cell(row=row_num+1,column=col_num+1).value = row['SERVICE']
            self.ws.cell(row=row_num+1,column=col_num+2).value = row['METHOD']
            self.ws.cell(row=row_num+1,column=col_num+3).value = row['PATH']
            if 'DESTINATION' in row:
                self.__insert(row['DESTINATION'],col_num+3)
                self.ws.merge_cells(start_row=row_num+1,start_column=col_num+1,end_row=self.ws.max_row,end_column=col_num+1)
                self.ws.merge_cells(start_row=row_num+1,start_column=col_num+2,end_row=self.ws.max_row,end_column=col_num+2)
                self.ws.merge_cells(start_row=row_num+1,start_column=col_num+3,end_row=self.ws.max_row,end_column=col_num+3)
    def __insert2(self,rows,col_num):
        for row in rows:
            row_num = self.ws.max_row
            if 'DESTINATION' in row:
                self.__insert2(row['DESTINATION'],col_num+3)
            self.ws.cell(row=row_num+1,column=col_num+1).value = row['SERVICE']
            self.ws.cell(row=row_num+1,column=col_num+2).value = row['METHOD']
            self.ws.cell(row=row_num+1,column=col_num+3).value = row['PATH']
            self.ws.merge_cells(start_row=row_num+1,start_column=col_num+1,end_row=self.ws.max_row,end_column=col_num+1)
            self.ws.merge_cells(start_row=row_num+1,start_column=col_num+2,end_row=self.ws.max_row,end_column=col_num+2)
            self.ws.merge_cells(start_row=row_num+1,start_column=col_num+3,end_row=self.ws.max_row,end_column=col_num+3)

    def saveSpecial(self,payload,level=2):
        self.__makeDir()
        self.ws.cell(row=1,column=1).value=''
        self.ws.cell(row=2,column=1).value=''
        col_num=0
        for pay in payload:
            row_num = self.ws.max_row
            if 'DESTINATION' in pay:
                self.__insert2(pay['DESTINATION'],3)
            self.ws.cell(row=row_num+1,column=col_num+1).value = pay['SERVICE']
            self.ws.cell(row=row_num+1,column=col_num+2).value = pay['METHOD']
            self.ws.cell(row=row_num+1,column=col_num+3).value = pay['PATH']
            self.ws.merge_cells(start_row=row_num+1,start_column=col_num+1,end_row=self.ws.max_row,end_column=col_num+1)
            self.ws.merge_cells(start_row=row_num+1,start_column=col_num+2,end_row=self.ws.max_row,end_column=col_num+2)
            self.ws.merge_cells(start_row=row_num+1,start_column=col_num+3,end_row=self.ws.max_row,end_column=col_num+3)
        
        self.__saveHeader(int(self.ws.max_column/3))
        # for cell in self.ws["1:1"]:
        #     cell.font=self.headerFont
        # for row in rows:
        #     self.ws.append(row)
        # self.ws.merge_cells('B'+str(self.ws.max_row-size)+':B'+str(self.ws.max_row))
    

        




        

                